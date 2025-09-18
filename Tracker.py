import os
import json
import logging
import requests
from dataclasses import dataclass
from bs4 import BeautifulSoup
from typing import List, Tuple, Optional, Dict
import openpyxl

# Set up logging
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# ---------------------- Configuration ----------------------

USERNAME = ""#USERNAME
PASSWORD = ""#PASSWORD
BASE_URL = ""#PATH TO THE SITE ORDERS
EXCEL_PATH = os.path.join(os.getcwd(), "Simple_Tracker.xlsx") #HARDCODED LOCALLY FILE
USER_ID = 52

# ---------------------- Data Models ------------------------

@dataclass
class OrderData:
    date_added: str
    status: str
    payment_left: int
    receipt_missing: bool
    interruption: str
    type: str
    customer_name: str
    telephone: str
    priority: int
    item_count: int
    purchase_info_text: str
    purchase_num: int

@dataclass
class PurchaseData:
    purchase_num: str
    remark: str
    status: str
    order_type: str
    priority: str
    shipment_location: str
    active: str = "INACTIVE"

# ---------------------- TMS Client -------------------------

class TMSClient:
    def __init__(self, username: str, password: str):
        self.session = requests.Session()
        self.token = self.login(username, password)

    def login(self, username: str, password: str) -> str:
        logging.info("Logging in to TMS...")
        url = f"{BASE_URL}/index.php?route=common/login"
        data = {'username': username, 'password': password}
        response = self.session.post(url, data=data)
        response.raise_for_status()
        token = response.text.split("token=")[1].split('"')[0]
        logging.info("Login successful.")
        return token

    def get_order_info(self, order_id: str) -> dict:
        url = f"{BASE_URL}/index.php?route=sale/order_new/get"
        params = {"order_id": order_id, "token": self.token}
        logging.info(f"Fetching order {order_id}")
        response = self.session.get(url, params=params)
        response.raise_for_status()
        return response.json()

    def get_purchase_info(self, purchase_id: str) -> str:
        url = f"{BASE_URL}/index.php?route=sale/purchase/edit&purchase_id={purchase_id}&token={self.token}"
        response = self.session.get(url)
        return response.content

    def get_all_user_orders(self) -> List[str]:
        url = f"{BASE_URL}/index.php"
        params = {
            'route': 'sale/order',
            'token': self.token,
            'filter_user_id': str(USER_ID),
            'filter_created_priority': '0',
            'filter_payment_priority': '0',
            'filter_verified_priority': '0',
        }
        response = self.session.get(url, params=params)
        return self._parse_order_ids(response.content)

    @staticmethod
    def _parse_order_ids(payload: bytes) -> List[str]:
        soup = BeautifulSoup(payload, "html.parser")
        orders = soup.find_all('td', class_="text-left")
        return [td.text.strip() for td in orders if td.text.strip().isnumeric()]

# ---------------------- Excel Handler -------------------------

class ExcelHandler:
    def __init__(self, path: str):
        self.path = path
        self.workbook = openpyxl.load_workbook(self.path)
        self.sheet = self.workbook.active

    def get_existing_orders(self) -> List[str]:
        orders = []
        row = 2
        while self.sheet.cell(row=row, column=1).value:
            orders.append(str(self.sheet.cell(row=row, column=1).value))
            row += 1
        return orders

    def get_orders_to_update(self) -> Dict[str, int]:
        update_list = {}
        row = 2
        while self.sheet.cell(row=row, column=1).value:
            if self.sheet.cell(row=row, column=19).value:
                order_id = str(self.sheet.cell(row=row, column=1).value)
                update_list[order_id] = row
            row += 1
        return update_list

    def insert_new_orders(self, orders: List[str]):
        existing = self.get_existing_orders()
        row = len(existing) + 2
        for order in orders:
            if order not in existing:
                self.sheet.cell(row=row, column=1).value = order
                row += 1

    def update_order(self, row: int, order_data: OrderData, purchase_data: Optional[PurchaseData]):
        self.sheet[f'B{row}'].value = order_data.date_added
        self.sheet[f'D{row}'].value = order_data.status
        self.sheet[f'C{row}'].value = order_data.interruption
        self.sheet[f'E{row}'].value = order_data.type
        self.sheet[f'F{row}'].value = order_data.customer_name
        self.sheet[f'G{row}'].value = order_data.telephone
        self.sheet[f'J{row}'].value = order_data.priority
        self.sheet[f'I{row}'].value = order_data.item_count
        self.sheet[f'K{row}'].value = order_data.purchase_info_text
        self.sheet[f'H{row}'].value = order_data.payment_left
        self.sheet[f'T{row}'].value = "UPLOAD_RECEIPT!" if order_data.receipt_missing else "OK!"

        if purchase_data:
            self.sheet[f'L{row}'].value = purchase_data.purchase_num
            self.sheet[f'M{row}'].value = purchase_data.remark
            self.sheet[f'N{row}'].value = purchase_data.status
            self.sheet[f'O{row}'].value = purchase_data.order_type
            self.sheet[f'P{row}'].value = purchase_data.priority
            self.sheet[f'Q{row}'].value = purchase_data.shipment_location
            self.sheet[f'R{row}'].value = purchase_data.active

        if self.sheet[f'D{row}'].value in ['Waiting for pickup', 'Shipped', 'Canceled', 'Completed']:
            self.sheet[f'S{row}'].value = False
        if purchase_data and purchase_data.status in ['CLOSED (CONFIRMED)', 'CLOSED (BY STOCK)']:
            self.sheet[f'S{row}'].value = False

    def save(self):
        self.workbook.save(self.path)

# ---------------------- Order Processing Logic -------------------------

class OrderProcessor:
    @staticmethod
    def parse_order(data: dict) -> OrderData:
        order = data['data']['order']
        status_map = {
            0: "CHECK?", 1: 'New Order', 3: 'Approved for operation',
            19: 'Shipped', 20: 'Waiting for pickup', 17: 'Completed',
            5: 'On Hold', 7: 'Canceled'
        }
        payment_left, receipt_missing = OrderProcessor.check_payment(order['charge_history'], float(order['totals']['total']['value']))
        purchase_info_text, purchase_num = OrderProcessor.check_purchase(order['order_products'])
        return OrderData(
            date_added=order['date_added'],
            status=status_map.get(order['order_status_id'], "Unknown"),
            payment_left=payment_left,
            receipt_missing=receipt_missing,
            interruption=OrderProcessor.check_interrupts(order['process_log']),
            type=OrderProcessor.get_order_type(order['order_type_id']),
            customer_name=f"{order['firstname']} {order['lastname']}",
            telephone=order['telephone'],
            priority=order['priority_id'],
            item_count=len(order['order_products']),
            purchase_info_text=purchase_info_text,
            purchase_num=purchase_num
        )

    @staticmethod
    def check_interrupts(logs: list) -> str:
        return "OK!" if all(log['user'] == "Leon Pechr" for log in logs) else "Check Interrupt!"

    @staticmethod
    def get_order_type(type_id: int) -> str:
        return {0: "NOT SET", 1: "Computer", 2: "Components"}.get(type_id, "Unknown")

    @staticmethod
    def check_payment(history: list, total: float) -> Tuple[int, bool]:
        remaining = total
        needs_receipt = False
        for charge in history:
            if charge.get('success'):
                remaining -= charge.get('total', 0)
                if not charge.get('priority_id') and charge.get('type') != "bank_transfer":
                    needs_receipt = True
        return int(max(0, remaining)), needs_receipt

    @staticmethod
    def check_purchase(products: list) -> Tuple[str, int]:
        try:
            claris = products[0]['number_order_claris']
        except IndexError:
            return "INDEX ERROR", -1
        if not claris:
            return "NO PURCHASE", -1
        for item in products:
            if item['number_order_claris'] != claris and item['name'] != 'TMS - לאון':
                return "CHECK PURCHASE DIFF", claris
        return "Purchase OK", claris

    @staticmethod
    def parse_purchase(raw_html: str) -> Optional[PurchaseData]:
        try:
            soup = BeautifulSoup(raw_html, 'html.parser')
            options = soup.find_all('option', selected=True)
            purchase_num = soup.find_all('h4')[1].text.split("PT")[1].strip()
            prio = soup.find('input', {'id': 'input-priority-number'})['value']
            remark = soup.body.findAll('input', attrs={'class': 'form-control', 'name': 'nickname'})[0]['value']

            status_map = {'1': 'DO NOT', '2': 'CLOSED (CONFIRMED)', '3': 'CLOSED (BY STOCK)'}
            type_map = {'0': 'NOT SELECTED', '1': 'Components', '2': 'Computer', '3': 'OUT'}
            ship_map = {
                '0': 'NOT SELECTED', 'Ntrn': 'Netanya', 'Htrn': 'Holon', 'Eman': 'EILAT',
                'Ndel': 'DELIVERY', 'Ttrn': 'Tel Aviv', 'Ftrn': 'Haifa', 'Btrn': 'Beer Sheva', 'Atrn': 'Ashdod'
            }

            return PurchaseData(
                purchase_num=purchase_num,
                remark=remark,
                status=status_map[options[0]['value']],
                order_type=type_map[options[1]['value']],
                priority=prio,
                shipment_location=ship_map[options[2]['value']]
            )
        except Exception as e:
            logging.error(f"Failed to parse purchase: {e}")
            return None

# ---------------------- Main -------------------------

def main():
    #INIT
    client = TMSClient(USERNAME, PASSWORD)
    excel = ExcelHandler(EXCEL_PATH)

    #Fetch and insert new orders
    all_orders = client.get_all_user_orders()
    existing_orders = excel.get_existing_orders()
    new_orders = [o for o in all_orders if o not in existing_orders]
    excel.insert_new_orders(new_orders)
    excel.save()

    #Update orders information
    to_update = excel.get_orders_to_update()
    for order_id, row in to_update.items():
        order_raw = client.get_order_info(order_id)
        order_data = OrderProcessor.parse_order(order_raw)

        purchase_data = None
        if order_data.purchase_num != -1:
            purchase_raw = client.get_purchase_info(order_data.purchase_num)
            purchase_data = OrderProcessor.parse_purchase(purchase_raw)

        excel.update_order(row=row, order_data=order_data, purchase_data=purchase_data)

    excel.save()

if __name__ == "__main__":
    main()